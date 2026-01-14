from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from datetime import datetime, timedelta
from pandas import ExcelFile, read_excel, isna,DataFrame
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from pathlib import Path


debug = False

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException,
    StaleElementReferenceException,
    ElementClickInterceptedException,
    ElementNotInteractableException,
)

def stable_click(driver, locator, logger, timeout=10, retries=3, post_wait=None):
    """
    locator example: (By.ID, "ctl00_MainContent_cmdNewClaim")
    post_wait: optional function(driver) -> True when click succeeded
    """
    last_err = None

    for attempt in range(1, retries + 1):
        try:
            el = WebDriverWait(driver, timeout).until(EC.presence_of_element_located(locator))

            # scroll to center (prevents header/overlay issues)
            driver.execute_script(
                "arguments[0].scrollIntoView({block:'center', inline:'center'});", el
            )

            # now wait for clickable (after scroll)
            el = WebDriverWait(driver, timeout).until(EC.element_to_be_clickable(locator))

            try:
                el.click()
            except (ElementClickInterceptedException, ElementNotInteractableException):
                logger.warning(f"Normal click failed, trying JS click (attempt {attempt})")
                driver.execute_script("arguments[0].click();", el)

            if post_wait:
                WebDriverWait(driver, timeout).until(lambda d: post_wait(d))

            return  # success

        except (StaleElementReferenceException, ElementClickInterceptedException,
                ElementNotInteractableException, TimeoutException) as e:
            last_err = e
            logger.warning(f"stable_click attempt {attempt} failed: {type(e).__name__}: {e}")

    raise last_err



def set_up_driver(link):
    # Set up WebDriver
    driver = webdriver.Chrome()
    time.sleep(1)
    driver.maximize_window()
    driver.get(link)
    return driver


def set_up_full_log_in(link, name, password, verification):
    driver = set_up_driver(link)

    # Wait for login fields
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "ctl00_MainContent_txtUser")))

    # Enter credentials
    driver.find_element(By.ID, "ctl00_MainContent_txtUser").send_keys(name)
    driver.find_element(By.ID, "ctl00_MainContent_txtPass").send_keys(password)
    driver.find_element(By.ID, "ctl00_MainContent_ExtraIdAnswer").send_keys(verification)

    # Click login button
    driver.find_element(By.ID, "ctl00_MainButtons_cmdOK").click()

    return driver


import os

def find_file_with_number(base_path, extracted_number):
    for root, dirs, files in os.walk(base_path):
        for file in files:
            if str(extracted_number) in file:
                file_path = os.path.join(root, file)
                return os.path.normpath(file_path)  # Normalize to system default
    return None



def process_excel(file_path,config, base_path="/"):
    customers = []
    print("working on:")

    try:
        # Load the Excel file with specific column types
        with ExcelFile(file_path, engine='openpyxl') as xls:
            df = read_excel(xls, dtype={2: str, 4: str})  # Read with column types

        for index, row in df.iterrows():
            # Check if column A (index 0) is empty
            if isna(row[config.id_col]):  # Check if column A (first column) is empty
                break  # Exit the loop

            # Get values from the row
            id_value = row[config.id_col]  # ID from column C (index 2)
            date_value = row[config.date_col]  # Date from column D (index 3)
            file_name = row[config.receipt_col]  # File name from column E (index 4)
            first_name = row[config.first_name_col]
            last_name = row[config.last_name_col]
            if config.model == "macabi":
                referral = ""
                need_new_referral = False
            else:  # clalit
                if  isna(row[config.new_approval_file_col]):
                    referral = ""
                    need_new_referral = False
                else:
                    referral = str(row[config.new_approval_file_col])
                    need_new_referral = referral != ""

                if base_path != "/"  and isinstance(file_name, str) and file_name.strip() != "":
                    match = re.search(r"\d{4,}", file_name)
                    if match:
                        extracted_number = match.group()
                        print("match", extracted_number)
                        file_name = find_file_with_number(base_path, extracted_number)

            # Convert date string to datetime object if needed
            if isinstance(date_value, str):
                date_value = datetime.strptime(date_value, '%Y-%m-%d')  # Adjust format if needed

            # Extract day, month, and year
            customers.append({
                "row": index + 2,  # Adding row number (1-based)
                "id": id_value,
                "day": date_value.day,
                "month": date_value.month,
                "year": date_value.year,
                "date": date_value,
                "file": file_name,
                "rows": [index + 2],  # Initialize rows list with the first occurrence
                "first_name": first_name,
                "last_name": last_name,
                "need_referral": need_new_referral,
                "referral" : referral,
                "write_to_excel": {config.first_name_col : first_name,
                                   config.last_name_col: last_name,
                                   config.id_col: id_value,
                                   config.date_col: date_value,
                                   config.did_reported_col: "X",
                                   config.error_col: "",
                                   }
            })
            if config.model == "clalit" and len(customers) > 0:
                customers[-1]["write_to_excel"][config.receipt_col] = "x"

            print(
                f"           Row: {index + 2}, ID: {id_value}, Date: {date_value.day}-{date_value.month}-{date_value.year}, file: {file_name}")

    except FileNotFoundError:
        print(f"Error: The file '{file_path}' was not found.")
    except PermissionError:
        print(f"Error: Permission denied. Close '{file_path}' if it's open.")
    except Exception as e:
        print(f"An error occurred: {e}")

    try:
        # find if there r duplicates base of id and name
        for i in range(len(customers)):
            for j in range(i + 1, len(customers)):
                # if there is the same customer with the same id and date
                if customers[i]["id"] == customers[j]["id"] and customers[i]["date"] == customers[j]["date"]:
                    # push the dup one day later
                    new_date = customers[j]["date"] + timedelta(days=1)
                    # if the month is different, push it to the previous day
                    if customers[i]["month"] != new_date.month :
                        new_date =  customers[j]["date"] - timedelta(days=1)
                    customers[j]["date"] = new_date
                    customers[j]["day"] = new_date.day
    except Exception as e:
        print(f"An error occurred while looking for duplicates: {e}")

    return customers


def get_unique_customers(customer_list):
    unique_customers = {}

    for customer in customer_list:
        customer_id = customer["id"]

        if customer_id in unique_customers:
            unique_customers[customer_id]["rows"].append(customer["row"])  # Append row number
        else:
            unique_customers[customer_id] = customer  # First occurrence

    return list(unique_customers.values())


def write_to_excel(file_path, row, col, txt):
    col = col + 1  # Adjust column index for openpyxl (1-based)
    try:
        # Load the existing Excel file
        wb = load_workbook(file_path)
        sheet = wb.active  # Get the active sheet
        # Write the text to the specified cell
        cell = sheet.cell(row=row, column=col)
        cell.value = txt


        # Center align the text in the cell
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if debug:
            print("STEP 6")
        # Save the modified file
        wb.save(file_path)
        wb.close()  # Close the workbook when you're done with it

        print(f"Text '{txt}' written to row {row}, column {col} (centered) in '{file_path}'")

    except FileNotFoundError:
        print(f"Error: The file '{file_path}' was not found.")
    except Exception as e:
        print(f"An error occurred: {repr(e)}")

def _excel_safe_value(v):
    if v is None:
        return ""
    # Selenium WebElement has these attrs; avoid importing selenium here if you want
    if hasattr(v, "tag_name") and hasattr(v, "get_attribute"):
        # choose what makes sense for your element:
        return v.text or v.get_attribute("value") or v.get_attribute("innerText") or ""
    # openpyxl supports: str, int, float, bool, datetime/date, None
    if isinstance(v, (str, int, float, bool)):
        return v
    # fallback
    return str(v)

def write_many_to_excel(file_path, writes):
    wb = load_workbook(file_path)
    sheet = wb.active

    for row, col, txt in writes:
        sheet.cell(row=row, column=col + 1).value = _excel_safe_value(txt)

    wb.save(file_path)

def update_customer_writing(customer, cols, texts):
    for i in range(len(cols)):
        customer["write_to_excel"][cols[i]] = texts[i]


def write_customer_to_excel(file_path, customer):
    row = customer["row"]

    writes = [
        (row, col, value)
        for col, value in customer["write_to_excel"].items()
    ]

    write_many_to_excel(file_path, writes)

def write_customer_to_excel_few_rows(customer,cols,output_xl_path):
    rows = customer["rows"]
    writes = []
    write_map = customer.get("write_to_excel", {})
    for r in rows:
        for c in cols:
            if c in write_map:
                writes.append((r, c,write_map[c]))

    if writes:
        write_many_to_excel(output_xl_path, writes)




def clear_col(file_path, col, end_of_col):
    for i in range(2, end_of_col + 2):
        write_to_excel(file_path, i, col, "")


def extract_date(alert_content):
    """
    Extracts the first date in the format XX/XX/XXXX from a given string.

    :param alert_content: The input element to search for a date.
    :return: The extracted date as a string if found, otherwise None.
    """

    alert_txt = alert_content.text
    print("alert txt = ", alert_txt)

    print(alert_txt)
    date_pattern = r"\b\d{2}/\d{2}/\d{4}\b"
    match = re.search(date_pattern, alert_txt)
    return match.group() if match else None


def copy_headers_by_index(input_xl_path, output_xl_path, header_indexes):
    """
    Creates an output Excel file and copies selected column headers
    from the input Excel by column indexes (0-based).

    :param input_xl_path: path to input Excel file
    :param output_xl_path: path to output Excel file (will be created)
    :param header_indexes: list of column indexes (0-based)
    """

    input_xl_path = Path(input_xl_path)
    output_xl_path = Path(output_xl_path)

    if not input_xl_path.exists():
        raise FileNotFoundError(f"Input Excel not found: {input_xl_path}")

    # Ensure output directory exists
    output_xl_path.parent.mkdir(parents=True, exist_ok=True)

    # Read only headers (no data)
    df = read_excel(input_xl_path, nrows=0)

    # Validate indexes
    max_index = len(df.columns) - 1
    invalid = [i for i in header_indexes if i < 0 or i > max_index]
    if invalid:
        raise IndexError(f"Invalid column indexes: {invalid}")

    # Extract selected headers
    selected_headers = [df.columns[i] for i in header_indexes]

    # Create empty DataFrame with selected headers
    out_df = DataFrame(columns=selected_headers)

    # Create the Excel file
    out_df.to_excel(output_xl_path, index=False)